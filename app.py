from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import requests
from bs4 import BeautifulSoup
import time
import traceback
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import urllib3
import os
import uuid
import gc

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)
CORS(app)

# Absolute paths — works on local and Railway/Render cloud
IS_CLOUD   = os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RENDER")
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = "/tmp/seo_uploads" if IS_CLOUD else os.path.join(BASE_DIR, "uploads")
REPORT_DIR = "/tmp/seo_reports" if IS_CLOUD else os.path.join(BASE_DIR, "reports")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(REPORT_DIR, exist_ok=True)

print(f"[INIT] Upload dir : {UPLOAD_DIR}")
print(f"[INIT] Report dir : {REPORT_DIR}")


# ── Exact same clean_text from your original script ───────────────────────────
def clean_text(text):
    if text is None or pd.isna(text):
        return ""
    return " ".join(str(text).replace("\n", " ").replace("\r", " ").split()).strip()


# ─────────────────────────────────────────────────────────────────────────────
#  ROUTE 1 — /upload
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/upload", methods=["POST"])
def upload():
    try:
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file provided"}), 400

        filename = f"{uuid.uuid4().hex}_{f.filename}"
        filepath = os.path.join(UPLOAD_DIR, filename)
        f.save(filepath)
        print(f"[UPLOAD] Saved: {filepath}")

        xl     = pd.ExcelFile(filepath)
        sheets = xl.sheet_names
        print(f"[UPLOAD] Sheets: {sheets}")
        return jsonify({"filepath": filepath, "sheets": sheets})

    except Exception:
        traceback.print_exc()
        return jsonify({"error": traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  ROUTE 2 — /get-columns
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/get-columns", methods=["POST"])
def get_columns():
    try:
        body       = request.get_json()
        filepath   = body.get("filepath", "")
        sheet_name = body.get("sheet_name", "")
        header_row = int(body.get("header_row", 1))

        print(f"[COLS] sheet={sheet_name}  header_row={header_row}")

        if not os.path.exists(filepath):
            return jsonify({"error": "Uploaded file not found. Please re-upload."}), 400

        # Exact same read as your script: header=1 means header_row=2 in UI
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row - 1)
        df.columns = df.columns.astype(str).str.strip()
        columns = df.columns.tolist()
        print(f"[COLS] {len(columns)} columns")
        return jsonify({"columns": columns})

    except Exception:
        traceback.print_exc()
        return jsonify({"error": traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  ROUTE 3 — /run-test
#  Core loop = your original script, character for character
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/run-test", methods=["POST"])
def run_test():
    try:
        body       = request.get_json()
        filepath   = body.get("filepath", "")
        sheet_name = body.get("sheet_name", "")
        header_row = int(body.get("header_row", 1))
        url_col    = body.get("url_col")
        h1_col     = body.get("h1_col")
        title_col  = body.get("title_col")
        meta_col   = body.get("meta_col")

        print(f"[RUN] sheet={sheet_name}  header_row={header_row}")
        print(f"[RUN] url={url_col}  h1={h1_col}  title={title_col}  meta={meta_col}")

        if not os.path.exists(filepath):
            return jsonify({"error": "Uploaded file not found. Please re-upload."}), 400
        if not url_col:
            return jsonify({"error": "URL column is required."}), 400

        # ── Read sheet — identical to your script ─────────────────────────────
        df = pd.read_excel(filepath, sheet_name=sheet_name, header=header_row - 1)
        df.columns = df.columns.astype(str).str.strip()

        # ── Keep only the mapped columns ──────────────────────────────────────
        required_columns = [c for c in [url_col, h1_col, title_col, meta_col] if c and c in df.columns]
        df = df[required_columns]

        # ── Rename to fixed internal names (same as your script) ──────────────
        rename_map = {}
        if url_col   and url_col   in required_columns: rename_map[url_col]   = "URL"
        if h1_col    and h1_col    in required_columns: rename_map[h1_col]    = "Expected_H1"
        if title_col and title_col in required_columns: rename_map[title_col] = "Expected_Title"
        if meta_col  and meta_col  in required_columns: rename_map[meta_col]  = "Expected_Meta"
        df = df.rename(columns=rename_map)

        # ── Ensure all expected columns exist even if not mapped ──────────────
        for col in ["URL", "Expected_H1", "Expected_Title", "Expected_Meta"]:
            if col not in df.columns:
                df[col] = ""

        # ── Remove empty URLs — identical to your script ──────────────────────
        df = df[df["URL"].notna()].reset_index(drop=True)
        df["URL"] = df["URL"].astype(str).str.strip()
        df = df[df["URL"].str.lower() != "nan"].reset_index(drop=True)

        # ── Result columns — identical to your script ─────────────────────────
        df["Status_Code"]   = ""
        df["Actual_H1"]     = ""
        df["H1_Count"]      = 0
        df["H1_Result"]     = ""
        df["Actual_Title"]  = ""
        df["Title_Result"]  = ""
        df["Actual_Meta"]   = ""
        df["Meta_Result"]   = ""
        # Force H1_Count to plain Python int column so .at[] assignment never fails
        df["H1_Count"] = df["H1_Count"].astype(object)

        print(f"[RUN] {len(df)} valid URL rows to process")

        # ── Per-request fetching to minimize memory on free tier ─────────────
        HEADERS = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}

        # ── Main loop — YOUR EXACT LOGIC, zero changes ────────────────────────
        issues = []
        passed = 0
        failed = 0

        for index, row in df.iterrows():

            url = row["URL"]

            expected_h1    = clean_text(row["Expected_H1"])
            expected_title = clean_text(row["Expected_Title"])
            expected_meta  = clean_text(row["Expected_Meta"])

            print("Checking:", url)

            try:
                response    = requests.get(url, timeout=10, verify=False, headers=HEADERS, stream=False)
                status_code = response.status_code

                # Grab text then immediately close connection to free memory
                html_text = response.text
                response.close()
                del response

                soup = BeautifulSoup(html_text, "html.parser")

                # H1
                h1_tags  = soup.find_all("h1")
                h1_count = len(h1_tags)
                actual_h1 = clean_text(h1_tags[0].get_text()) if h1_count > 0 else ""

                # Title
                actual_title = clean_text(soup.title.get_text()) if soup.title else ""

                # Meta description
                meta_tag = soup.find("meta", attrs={"name": "description"})
                if not meta_tag:
                    meta_tag = soup.find("meta", attrs={"property": "og:description"})
                actual_meta = clean_text(meta_tag.get("content")) if meta_tag and meta_tag.get("content") else ""

                # Free soup from memory immediately after extracting data
                soup.decompose()
                del soup, html_text

            except Exception:
                print("Error loading:", url)
                status_code = "ERROR"
                actual_h1   = ""
                actual_title = ""
                actual_meta  = ""
                h1_count     = 0

            # Save extracted values — cast to avoid pandas dtype mismatch
            df.at[index, "Status_Code"]  = str(status_code)
            df.at[index, "Actual_H1"]    = str(actual_h1)
            df.at[index, "H1_Count"]     = int(h1_count)   # keep as int
            df.at[index, "Actual_Title"] = str(actual_title)
            df.at[index, "Actual_Meta"]  = str(actual_meta)

            # ---------------- H1 Result — YOUR EXACT LOGIC ----------------
            if expected_h1 == "":
                h1_result = "NO EXPECTED H1"
            elif actual_h1 == "":
                h1_result = "H1 MISSING"
            elif h1_count > 1:
                h1_result = "MULTIPLE H1"
            elif expected_h1.lower() == actual_h1.lower():
                h1_result = "PASS"
            else:
                h1_result = "FAIL"

            # ---------------- Title Result — YOUR EXACT LOGIC ----------------
            if expected_title == "":
                title_result = "NO EXPECTED TITLE"
            elif actual_title == "":
                title_result = "TITLE MISSING"
            elif expected_title.lower() == actual_title.lower():
                title_result = "PASS"
            else:
                title_result = "FAIL"

            # ---------------- Meta Result — YOUR EXACT LOGIC ----------------
            if expected_meta == "":
                meta_result = "NO EXPECTED META"
            elif actual_meta == "":
                meta_result = "META MISSING"
            elif expected_meta.lower() == actual_meta.lower():
                meta_result = "PASS"
            else:
                meta_result = "FAIL"

            df.at[index, "H1_Result"]    = h1_result
            df.at[index, "Title_Result"] = title_result
            df.at[index, "Meta_Result"]  = meta_result

            # ── Build popup result entry ──────────────────────────────────────
            row_checks = []
            row_failed = False

            if h1_col:
                row_checks.append({"label": "H1",    "pass": h1_result    == "PASS", "result": h1_result,    "actual": actual_h1[:100],    "expected": expected_h1[:100]})
                if h1_result != "PASS": row_failed = True
            if title_col:
                row_checks.append({"label": "Title", "pass": title_result == "PASS", "result": title_result, "actual": actual_title[:100], "expected": expected_title[:100]})
                if title_result != "PASS": row_failed = True
            if meta_col:
                row_checks.append({"label": "Meta",  "pass": meta_result  == "PASS", "result": meta_result,  "actual": actual_meta[:100],  "expected": expected_meta[:100]})
                if meta_result != "PASS": row_failed = True
            if str(status_code) not in ("200", ""):
                row_checks.append({"label": "HTTP", "pass": False, "result": str(status_code), "actual": str(status_code), "expected": "200"})
                row_failed = True

            if row_failed:
                failed += 1
                issues.append({"url": url, "checks": row_checks})
            else:
                passed += 1

            # Delay to avoid blocking — identical to your script
            time.sleep(0.5)
            gc.collect()  # force free memory after each URL

        total = passed + failed
        print(f"\n[RUN] Done. total={total}  passed={passed}  failed={failed}")

        # ── Save output — identical to your script ────────────────────────────
        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"seo_results_{timestamp}.xlsx"
        report_path = os.path.join(REPORT_DIR, output_file)

        df.to_excel(report_path, index=False)

        # ── Highlight FAIL cells — YOUR EXACT LOGIC ───────────────────────────
        wb = load_workbook(report_path)
        ws = wb.active

        # Red    = FAIL
        # Orange = MISSING (H1 MISSING / TITLE MISSING / META MISSING)
        # Yellow = MULTIPLE H1
        fail_fill     = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        missing_fill  = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        multiple_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

        columns_to_check = ["H1_Result", "Title_Result", "Meta_Result"]
        col_map = {}

        for cell in ws[1]:
            if cell.value in columns_to_check:
                col_map[cell.value] = cell.column

        for row_num in range(2, ws.max_row + 1):
            for col in col_map.values():
                cell = ws.cell(row=row_num, column=col)
                val  = str(cell.value).strip().upper() if cell.value else ""
                if val == "FAIL":
                    cell.fill = fail_fill
                elif "MISSING" in val:
                    cell.fill = missing_fill
                elif val == "MULTIPLE H1":
                    cell.fill = multiple_fill

        wb.save(report_path)
        print(f"[RUN] Report saved: {report_path}")

        return jsonify({
            "summary":      {"total": total, "passed": passed, "failed": failed},
            "issues":       issues,
            "download_url": f"/download/{output_file}",
        })

    except Exception:
        traceback.print_exc()
        return jsonify({"error": traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  ROUTE 4 — /download/<filename>
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/download/<filename>")
def download(filename):
    path = os.path.join(REPORT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"error": "Report not found"}), 404
    return send_file(
        path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/")
def health():
    return jsonify({"status": "SEO QA API running", "version": "2.2"})


if __name__ == "__main__":
    app.run(debug=True, port=5000)
